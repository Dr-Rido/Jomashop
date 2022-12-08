# coding=utf-8
import re
import time
import json
import scrapy
import openpyxl
from scrapy.cmdline import execute
from jomashop_com.items import JomashopComItem
from math import ceil


class JomashopSpider(scrapy.Spider):
    name = 'Jomashop'
    allowed_domains = ['jomashop.com']
    start_urls = ['https://www.jomashop.com/']
    count = 0
    
    def start_requests(self):
        test_path = r"D:\PythonProject\jomashop_com\jomashop.xlsx"
        online_path = '/home/jomashop_com/jomashop.xlsx'
        wb = openpyxl.load_workbook(test_path)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            url = ws.cell(r, 1).value
            one, two, three = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
            meta = {'start_url': url, "category": one, "sub_category": two, "third_category": three if three else ''}
            meta['page_num'] = 1
            yield scrapy.Request(url=url, callback=self.parse, meta=meta)

    def parse(self, response, **kwargs):
        """请求json包"""
        meta = response.meta
        # print(f'当前所在页数:{num}')
        id = int(re.findall("data-model-id=\"(\d+)\"", response.text)[0])
        gender, subtype = self.parse_params(response.url)
        API = self.get_API(gender=gender, subtype=subtype, id=id, page=meta['page_num'])
        # print(API)
        yield scrapy.Request(url=API, callback=self.parse_list, meta=meta)

    def parse_list(self, response):
        """解析json包"""
        meta = response.meta
        json_data = json.loads(response.text)
        product_list = json_data['data']['products']['items']
        meta['sub_category'] = json_data['data']['categoryList'][0]['name']
        meta['third_category'] = [i['options'][1]['label'] for i in json_data['data']['products']['aggregations'] if i['attribute_code'] == 'subtype'][0]
        for product in product_list:
            meta['url'] = 'https://www.jomashop.com/' + product['url_key'] + '.html'
            meta['price_now'] = product['price_range']['minimum_price']['final_price']['value']
            meta['price_original'] = ''
            meta['discount'] = ''
            if product.get("msrp") and meta['price_now']:
                meta['price_original'] = product['msrp']
                meta['discount'] = str(round(float(meta['price_now']) / float(meta['price_original']), 2))
            meta['product_small_image'] = product['small_image']['sizes'][0]['url']
            meta['brand_name'] = product['brand_name']
            meta['title'] = product['name_wout_brand']
            meta['uuid'] = product['id']
            meta['urlKey'] = product['url_key']
            detail_api = self.detail_api(meta['urlKey'])
            yield scrapy.Request(url=detail_api, callback=self.parse_detail, meta=meta)

        # 翻页
        product_count = json_data['data']['products']['total_count']
        # print(product_count)
        max_page = ceil(product_count/60)
        if max_page >= 2:
            for page_num in range(2, max_page+1):
                meta['page_num'] = page_num
                yield scrapy.Request(url=meta['start_url'], callback=self.parse, meta=meta, dont_filter=True)

    def parse_detail(self, response):
        meta = response.meta
        item = JomashopComItem()
        self.count += 1
        json_data = json.loads(response.text)
        save_tag = True
        # 固定字段
        item['currency'] = '$'
        item['lang_country'] = 'en-us'
        item['source_id'] = 42
        item['add_time'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        item['update_time'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        item['delete_time'] = ''
        item['category'] = meta['category']
        item['sub_category'] = meta['sub_category']
        item['third_category'] = meta['third_category']
        item['uploaded'] = 0
        # item['subdivision_cat'] = meta['subdivision_cat']
        item['id'] = self.count
        # 爬取字段
        item['url'] = meta['url']
        item['title'] = meta['title']
        item['brand_name'] = meta['brand_name']
        item['price_now'] = str(meta['price_now']) if meta['price_now'] else ''
        item['price_original'] = str(meta['price_original'])
        item['discount'] = meta['discount']
        if not item['price_now']:
            try:
                item['price_now'] = str(json_data['data']['productDetail']['items'][0]['variants'][0]['product']['price_range']['minimum_price']['plp_price']['now_price'])
                item['price_original'] = str(json_data['data']['productDetail']['items'][0]['variants'][0]['product']['msrp'])
                item['discount'] = str(round(float(item['price_now']) / float(item['price_original']), 2))
            except:
                save_tag = False
                item['price_now'] = ''
                item['price_original'] = ''
                item['discount'] = ''
        if item['discount'] == '1.0':
            item['price_original'] = ''
            item['discount'] = ''
        # item['uuid'] = meta['uuid']
        item['uuid'] = json_data['data']['productDetail']['items'][0]['model_id']
        item['product_small_image'] = meta['product_small_image']
        try:
            item['product_big_images'] = [img['sizes'][1]['url'] for img in json_data['data']['productDetail']['items'][0]['media_gallery']]
            item['product_thumb_images'] = [img['sizes'][0]['url'] for img in json_data['data']['productDetail']['items'][0]['media_gallery']]
        except:
            item['product_big_images'] = []
            item['product_thumb_images'] = []
        item['product_details'] = ''
        item['image_urls'] = [item['product_small_image']] + item['product_big_images'] + ['product_thumb_images']
        item['images'] = ''
        # 出站信息
        item['out_url'] = meta['url']
        item['out_url_status'] = 1  # 自己,填1
        item['retailler_domain'] = meta['url'].split("/")[2]
        # 新增字段
        more_details_list = json_data['data']['productDetail']['items'][0]['moredetails']['more_details']
        Color_list = []
        Material_list = []
        for details in more_details_list:
            [Color_list.append(i['attribute_value']) for i in details['group_attributes'] if re.findall('.*Color.*', i['attribute_label'])]
            [Material_list.append(i['attribute_value']) for i in details['group_attributes'] if re.findall('.*Material.*', i['attribute_label'])]
        if Color_list:
            item['color'] = ','.join(Color_list)
        else:
            item['color'] = ''
        if Material_list:
            item['material'] = ''.join(Material_list)
        else:
            item['material'] = ''

        description_str = json_data['data']['productDetail']['items'][0]['description']['html']
        item['description'] = re.sub('<.*>', '', description_str)

        if save_tag:
            yield item

    def get_API(self, gender, subtype, id: int, page=1):
        """
        修改gender、subtype、id参数
        :param page:
        :param gender:
        :param subtype:
        :param id:
        :return:
        """
        str1 = 'https://www.jomashop.com/graphql?operationName=category&variables='
        str2 = '{"currentPage":1,"id":63,"onServer":true,"pageSize":60,"filter":{"category_id":{"eq":"63"},"gender":{"in":["Ladies"]},"subtype":{"in":["Swim"]}},"sort":{}}'
        str3 = '&extensions={"persistedQuery":{"version":1,"sha256Hash":"7f527cf964527a903b479b2de520f28573aed1a8204e00c29e285e75c46d5b3a"}}'
        str2 = str2.replace('true', '66666666666666666666')  # 字典里存在ture会报错
        dic = eval(str2)
        dic['currentPage'] = page
        dic['id'] = id
        dic['filter']['category_id']['eq'] = str(id)
        if type(gender) == str:
            dic['filter']['gender']['in'] = [gender]
        else:
            dic['filter']['gender']['in'] = gender
        dic['filter']['subtype']['in'] = [subtype]
        dic_str = json.dumps(dic)
        dic_str = dic_str.replace('66666666666666666666', 'true')
        api = str1 + dic_str + str3
        return api

    def detail_api(self, urlKey):
        """详情页API"""
        dic_str = '{"urlKey":"versace-sneakers-dsr611d-dna5pr-k0at","onServer":true}'
        end_str = '&extensions={"persistedQuery":{"version":1,"sha256Hash":"9ad1970a5262a81c75e35a7d599f1f6cc3f45d8b499794401d149530bdcc47cb"}}'
        head_str = 'https://www.jomashop.com/graphql?operationName=productDetail&variables='
        dic_str = dic_str.replace('true', '66666666666666666666')
        dic = eval(dic_str)
        dic['urlKey'] = urlKey
        dic_str = json.dumps(dic)
        dic_str = dic_str.replace('66666666666666666666', 'true')
        api = head_str + dic_str + end_str
        return api

    def off_(self, str):
        """用unquote会把%7C替掉不方便后面的替换"""
        if '%27' in str:
            str = str.replace('%27', '\'')
        if '%26+' in str:
            str = str.replace('%26+', '')
        if '%7C' in str:
            str = str.replace('%7C', '|')
        if '|' in str:
            str = str.split('|')
        if '+' in str:
            str = str.replace('+', ' ')
        return str

    def parse_params(self, url):
        p_list = url.split('?')
        params_str = p_list[1]
        if '&' in params_str:
            params_list = params_str.split('&')
        else:
            params_list = params_str
        if type(params_list) == list:
            params_list.sort()
            gender = re.findall("gender=(.*)", params_list[0])[0]
            subtype = re.findall("subtype=(.*)", params_list[1])[0]
        else:
            g = re.findall("gender=(.*)", params_list)
            s = re.findall("subtype=(.*)", params_list)
            gender = g[0] if g else ''
            subtype = s[0] if s else ''
        gender = self.off_(gender)
        subtype = self.off_(subtype)
        # print(gender, subtype)
        return gender, subtype


def start():
    execute("scrapy crawl Jomashop".split())


if __name__ == '__main__':
    start()
    